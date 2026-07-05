import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

DB = "db/nifty100.db"

OUTPUT = Path("output")
OUTPUT.mkdir(exist_ok=True)

OUTFILE = OUTPUT / "peer_comparison.xlsx"

conn = sqlite3.connect(DB)

peer_groups = pd.read_sql_query(
    """
    SELECT *
    FROM peer_groups
    """,
    conn,
)

companies = pd.read_sql_query(
    """
    SELECT id,
           company_name
    FROM companies
    """,
    conn,
)

ratios = pd.read_sql_query(
    """
    SELECT *
    FROM financial_ratios
    WHERE year = (
        SELECT MAX(year)
        FROM financial_ratios f
        WHERE f.company_id = financial_ratios.company_id
          AND f.year!='TTM'
    )
    """,
    conn,
)

percentiles = pd.read_sql_query(
    """
    SELECT *
    FROM peer_percentiles
    """,
    conn,
)

conn.close()

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GOLD = PatternFill("solid", fgColor="FFD966")
HEADER = PatternFill("solid", fgColor="D9EAD3")

metrics = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

with pd.ExcelWriter(OUTFILE, engine="openpyxl") as writer:

    for group in sorted(peer_groups["peer_group_name"].dropna().unique()):

        grp = peer_groups[
            peer_groups["peer_group_name"] == group
        ]

        benchmark = grp["benchmark_company"].iloc[0]

        df = (
            grp.merge(
                companies,
                left_on="company_id",
                right_on="id",
                how="left",
            )
            .merge(
                ratios,
                on="company_id",
                how="left",
            )
        )

        for metric in metrics:

            pct = percentiles[
                (percentiles.peer_group_name == group)
                &
                (percentiles.metric == metric)
            ][["company_id", "percentile_rank"]]

            pct = pct.rename(
                columns={
                    "percentile_rank":
                    metric + "_percentile"
                }
            )

            df = df.merge(
                pct,
                on="company_id",
                how="left",
            )

        export = df.drop(columns=["id"])

        sheet = group[:31]

        export.to_excel(
            writer,
            sheet_name=sheet,
            index=False,
        )

        ws = writer.book[sheet]

        for cell in ws[1]:
            cell.fill = HEADER
            cell.font = Font(bold=True)

        percentile_cols = []

        for c in range(1, ws.max_column + 1):

            name = ws.cell(row=1, column=c).value

            if name.endswith("_percentile"):
                percentile_cols.append(c)

            width = max(
                len(str(name)),
                14,
            )

            ws.column_dimensions[
                get_column_letter(c)
            ].width = width

        benchmark_row = None

        for r in range(2, ws.max_row + 1):

            if ws.cell(r, 1).value == benchmark:
                benchmark_row = r

            for c in percentile_cols:

                value = ws.cell(r, c).value
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    continue

                if value is None:
                    continue

                if value >= 0.75:
                    ws.cell(r, c).fill = GREEN

                elif value >= 0.25:
                    ws.cell(r, c).fill = YELLOW

                else:
                    ws.cell(r, c).fill = RED

        if benchmark_row:

            for c in range(1, ws.max_column + 1):
                ws.cell(
                    benchmark_row,
                    c,
                ).fill = GOLD

        median = ["Median", ""]

        for col in export.columns[2:]:

            if pd.api.types.is_numeric_dtype(export[col]):

                median.append(
                    export[col].median()
                )

            else:
                median.append("")

        ws.append(median)

print("Saved:", OUTFILE)
print(
    "Sheets:",
    len(peer_groups["peer_group_name"].unique()),
)