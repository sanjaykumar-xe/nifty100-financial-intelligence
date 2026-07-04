from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from src.screener.engine import load_config, load_latest_financial_data, apply_filters

OUTPUT = Path("output")
OUTPUT.mkdir(exist_ok=True)

EXPORT_FILE = OUTPUT / "screener_output.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def export_screeners():
    config = load_config()
    df = load_latest_financial_data()

    columns = [
        "company_id",
        "company_name",
        "broad_sector",
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "icr_label",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "asset_turnover",
        "sales",
        "net_profit",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "dividend_payout_ratio_pct",
        "market_cap_crore",
        "composite_quality_score",
    ]

    with pd.ExcelWriter(EXPORT_FILE, engine="openpyxl") as writer:
        for preset_name, filters in config["presets"].items():
            result = apply_filters(df, filters)

            sheet_name = preset_name[:31]
            export_cols = [col for col in columns if col in result.columns]

            result[export_cols].to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]

            for cell in ws[1]:
                cell.fill = HEADER
                cell.font = Font(bold=True)

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    print(f"Saved {EXPORT_FILE}")


if __name__ == "__main__":
    import pandas as pd
    export_screeners()